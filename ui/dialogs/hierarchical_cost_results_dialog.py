# ui/dialogs/hierarchical_cost_results_dialog.py
# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QTreeWidget, QTreeWidgetItem, QTabWidget, QWidget,
    QHeaderView, QFrame, QSplitter, QMessageBox, QFileDialog
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QColor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import (
    get_column_letter,
    column_index_from_string
)
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import pandas as pd

class HierarchicalCostResultsDialog(QDialog):
    """Dialog to display improved hierarchical project cost calculation results."""
    
    def __init__(self, cost_data, parent=None):
        super().__init__(parent)
        self.cost_data = cost_data
        self.df = pd.DataFrame(cost_data, columns=["Subtitle", 
                                                    "Province", 
                                                    "Description", 
                                                    "Target_Audience",
                                                    "Code",
                                                    "Unit", 
                                                    "Cost",
                                                    "Unit Cost (VND)", 
                                                    "Total Cost (VND)",
                                                    "Comment"])
        self.init_ui()
        
    def init_ui(self):
        """Initialize the UI components."""
        # Set dialog properties
        self.setWindowTitle("Hierarchical Project Cost Results")
        self.setMinimumSize(1240, 700)
        
        # Main layout
        main_layout = QVBoxLayout(self)
        
        # Header
        header_label = QLabel("Hierarchical Project Cost Calculation Results")
        header_font = QFont()
        header_font.setPointSize(16)
        header_font.setBold(True)
        header_label.setFont(header_font)
        header_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(header_label)
        
        # # Total cost display
        # total_cost = self.cost_data.get("total_cost", 0)
        # total_cost_label = QLabel(f"Total Project Cost: {total_cost:,.0f} VND")
        # total_cost_font = QFont()
        # total_cost_font.setPointSize(14)
        # total_cost_font.setBold(True)
        # total_cost_label.setFont(total_cost_font)
        # total_cost_label.setAlignment(Qt.AlignCenter)
        # total_cost_label.setStyleSheet("color: #2C3E50; background-color: #ECF0F1; padding: 10px; border-radius: 5px;")
        # main_layout.addWidget(total_cost_label)
        
        # Create tabs for provinces
        self.tabs = QTabWidget()
        
        # Add summary tab
        # summary_tab = self.create_summary_tab()
        # self.tabs.addTab(summary_tab, "Summary")
        
        # Add a tab for each province
        # provinces = self.cost_data.get("provinces", {})

        # for province, province_data in provinces.items():
        #     province_tab = self.create_province_tab(province, province_data)
        #     self.tabs.addTab(province_tab, province)
        
        # main_layout.addWidget(self.tabs)
        
        estimate_cost_tab = self.create_estimate_cost_tab()
        self.tabs.addTab(estimate_cost_tab, "Estimate Cost")

        main_layout.addWidget(self.tabs)

        # Buttons layout
        buttons_layout = QHBoxLayout()
        
        # Export button
        export_button = QPushButton("Export Results")
        export_button.clicked.connect(self.export_results)
        
        # Close button
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        
        # Add buttons to layout
        buttons_layout.addWidget(export_button)
        buttons_layout.addStretch()
        buttons_layout.addWidget(close_button)
        
        # Add buttons layout to main layout
        main_layout.addLayout(buttons_layout)
    
    def create_estimate_cost_tab(self):
        """Create a tab for a specific province."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Province total cost
        # total_cost = province_data.get("total_cost", 0)
        # total_label = QLabel(f"Total Cost for {province}: {total_cost:,.0f} VND")
        # total_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 10px;")
        # layout.addWidget(total_label)
        
        # Create tree widget for costs
        tree = QTreeWidget()
        tree.setHeaderLabels([
            "Subtitle / Component", 
            "Province", 
            "Target Audience", 
            "Unit",
            "Qty", 
            "Unit Cost (VND)", 
            "Total Cost (VND)",
            "Comment"
        ])
        # tree.setAlternatingRowColors(True)
        
        # Set column widths
        tree.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Subtitle / Component
        tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Province
        tree.header().setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Target Audience
        tree.header().setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Unit
        tree.header().setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Qty
        tree.header().setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Unit Cost
        tree.header().setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Total Cost
        tree.header().setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Comment
        
        # Set minimum column widths
        tree.setColumnWidth(1, 80)  
        tree.setColumnWidth(2, 80)  
        tree.setColumnWidth(3, 150)  
        tree.setColumnWidth(4, 100) 
        tree.setColumnWidth(5, 100)  
        tree.setColumnWidth(6, 100)  
        tree.setColumnWidth(7, 200)  
        
        # Add cost hierarchy
        self.add_cost_hierarchy(tree, self.cost_data)
        
        layout.addWidget(tree)
        
        # Expand top-level items
        tree.expandAll()
        
        # Enable custom tooltip handling
        tree.setMouseTracking(True)
        
        return widget
    
    def calculate_total_cost(self, titles=list(), province=""):
        if province:
            df = self.df[self.df["Province"] == province]
        else:
            df = self.df

        condition = True

        for i, title in enumerate(titles):
            condition &= df["Subtitle"].str.split(' / ').str[i].str.contains(title, case=False, na=False)

        df = df[condition]
        
        return df["Total Cost (VND)"].sum()

    def add_cost_hierarchy(self, tree, hierachy):
        subtitle = ""
        province_nodes = {}

        for i, row in enumerate(hierachy):
            subtitle = row[0]
            province_title = row[1]

            if "TRAVEL" in subtitle:
                a = ""

            subtitles = row[0].split(' / ')
            title_root = subtitles[0]

            root = tree.invisibleRootItem() # Root gốc
            current_root = None

            for i in range(root.childCount()):
                item_node = root.child(i)

                if item_node.text(0) == title_root:
                    current_root = item_node
                    break
            
            if not current_root:
                total_cost = self.calculate_total_cost(titles=[title_root])

                node = self.create_node(title_root, total_cost=total_cost)
                tree.addTopLevelItem(node)
                current_root = node

            province_key = (current_root, province_title)
            
            if province_key not in province_nodes:
                total_cost = self.calculate_total_cost(titles=[title_root], province=province_title)

                province_node = self.create_node(province_title, total_cost=total_cost)
                current_root.addChild(province_node)
                province_nodes[province_key] = province_node
            else:
                province_node = province_nodes[province_key]

            current_parent = province_node

            for title in subtitles[1:]:
                found = False
                for i in range(current_parent.childCount()):
                    child = current_parent.child(i)
                    if child.text(0) == title:
                        current_parent = child
                        found = True
                        break

                if not found:
                    new_node = self.create_node(title, total_cost=0)
                    current_parent.addChild(new_node)
                    current_parent = new_node
            
            child_node = self.create_child_node(row, current_parent)
            current_parent.addChild(child_node)

    def create_node(self, title, total_cost=0):
            
        item = QTreeWidgetItem([title])
        item.setTextAlignment(0, Qt.AlignLeft | Qt.AlignVCenter)

        if total_cost > 0:
            item.setText(6, f"{total_cost:,.0f}")
            item.setTextAlignment(6, Qt.AlignRight | Qt.AlignVCenter)

        # Set subtitle styling
        font = item.font(0)
        font.setBold(True)
        item.setFont(0, font)
        item.setFont(6, font)

        for i in range(item.columnCount()):
            item.setBackground(i, QColor("#F2F2F2"))

        return item

    def create_child_node(self, data, current_parent):
        item = QTreeWidgetItem()

        item.setText(0, data[2])                #Subtitle / Component
        item.setText(1, str(data[1]))           #Province
        item.setText(2, data[3])                #Target Audience
        item.setText(3, data[5] if data[5] != "0" else "-")                #Unit
        
        if data[0] in ["SUPERVISOR/ ASSISTANT", "QC", "DP"]:
            item.setText(4, f"{data[7]:,.2f}") #Qty
        else:               
            item.setText(4, f"{data[7]:,.0f}") #Qty     
            
        item.setText(5, f"{data[6]:,.0f}")       #Unit Cost (VND)
        item.setText(6, f"{data[8]:,.0f}")      #Total Cost (VND)
        item.setText(7, data[9]) #Comment

        item.setTextAlignment(1, Qt.AlignCenter)
        item.setTextAlignment(2, Qt.AlignCenter)
        item.setTextAlignment(3, Qt.AlignCenter)
        item.setTextAlignment(4, Qt.AlignRight | Qt.AlignVCenter)
        item.setTextAlignment(5, Qt.AlignRight | Qt.AlignVCenter)
        item.setTextAlignment(6, Qt.AlignRight | Qt.AlignVCenter)
        item.setTextAlignment(7, Qt.AlignLeft)

         # Set tooltips with formulas
        item.setToolTip(4, "Formular")  # Qty

        return item
    
    def export_results(self):
        """Export cost results to an Excel file."""
        # Open file save dialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results",
            "project_costs.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        # Ensure .xlsx extension
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Try to get project model from parent window
        project_model = None

        if hasattr(self.parent(), 'project_model'):
            project_model = self.parent().project_model
        
        self.create_estimate_cost_sheet(wb, project_model=project_model)

        # Create Project Information sheet
        # self.create_project_info_sheet(wb, project_model)
        
        # # Create Summary sheet
        # self.create_summary_sheet(wb, project_model)
        
        # # Create sheets for each province
        # provinces = self.cost_data.get("provinces", {})
        # for province, province_data in provinces.items():
        #     self.create_province_sheet(wb, province, province_data)
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Save the workbook
        try:
            wb.save(file_path)
            QMessageBox.information(
                self,
                "Export Successful",
                f"Results exported to {file_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to save file: {str(e)}"
            )

    def create_estimate_cost_sheet(self, wb, project_model=None):
        """Create and populate the Summary sheet with total costs across all provinces."""
        # Create sheet
        sheet = wb.create_sheet("Estimate Cost")
        
        # Set column widths for initial columns
        estimate_cost_sheet = {
            'table' : {
                'headers' : {
                    "Subtitle / Component" : {
                        "column_dimension" : "A",
                        "width" : 40,
                        "format": "",  # format Excel
                        "align": "left",   # align cell
                        "visible": False    
                    }, 
                    "Code" : {
                        'column_dimension' : 'B',
                        'width' : 15,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False
                    }, 
                    "Unit" : {
                        'column_dimension' : 'C',
                        'width' : 10,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False  
                    },
                    "Target Audience" : {
                        'column_dimension' : 'D',
                        'width' : 20,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False  
                    },
                    "Qty" : {
                        'column_dimension' : 'E',
                        'width' : 15,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False  
                    },
                    "Unit Cost (VND)" : {
                        'column_dimension' : 'F',
                        'width' : 20,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False  
                    },
                    "Total Cost (VND)" : {
                        'column_dimension' : 'G',
                        'width' : 20,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False
                    },
                    "Note" : {
                        'column_dimension' : 'H',
                        'width' : 20,
                        "format": "",  # format Excel
                        "align": "right",   # align cell
                        "visible": False  
                    }
                }
            }
        }

        row = 2

        column_dimensions = [properties.get('column_dimension') for properties in estimate_cost_sheet['table']['headers'].values()]

        cell = sheet.cell(row=row, column=1)
        cell.value = "ESTIMATE COST"
        cell.font = Font(bold=True, size=28)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        estimate_cost_sheet['table']['headers']
        sheet.merge_cells(f'{column_dimensions[0]}{row}:{column_dimensions[-1]}{row}')

        # Thông tin dự án
        row = 4

        project_mappings = {
            "project_name" : "Internal Job:",
            "internal_job" : "Project name:",
            "project_type" : "Method:"
        }

        for key, value in project_mappings.items():
            cell = sheet.cell(row=row, column=1)
            cell.value = value
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')

            cell = sheet.cell(row=row, column=2)
            cell.value = project_model.general.get(key, "")

            row += 1

        # Thông tin Sample Size theo từng thành phố
        row, column = 4, 4

        headers = ["Location", "Target Audience", "Sample", "Over sample", "Total"]

        for header in headers:
            cell = sheet.cell(row=row, column=column)
            cell.value = header

            column += 1

        row, column = 5, 4
        
        for province, target_audiences in project_model.samples.items():
            for key, target_audience in target_audiences.items():
                total = 0
                daily_interview_target = target_audience.get('target', {}).get('daily_interview_target', 0)
                target_for_interviewer = target_audience.get('target', {}).get('target_for_interviewer', 0) 

                for _col, header in enumerate(headers):
                    cell = sheet.cell(row=row, column=column + _col)
                    
                    if header == "Location":
                        cell.value = f"{province} - {target_audience.get('sample_type')}"
                    elif header == "Target Audience":
                        cell.value = target_audience.get('name')
                    elif header == "Sample":
                        cell.value = target_audience.get('sample_size')
                        total += cell.value
                    elif header == "Over sample":
                        cell.value = round(target_audience.get('sample_size') * target_audience.get('extra_rate') / 100, 0)
                        total += cell.value
                    else:
                        cell.value = total

                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                row += 1

        row = 9 if row < 9 else row

        cell = sheet.cell(row=row, column=5)
        cell.value = "Daily Interview Target:"

        cell = sheet.cell(row=row, column=6)
        cell.value = daily_interview_target

        cell = sheet.cell(row=row, column=5)
        cell.value = "Daily Interview Target:"

        cell = sheet.cell(row=row, column=6)
        cell.value = daily_interview_target

        
        for key, properties in estimate_cost_sheet['table']['headers'].items():
            column_dimension = properties.get('column_dimension')
            column_width = properties.get('width')

            sheet.column_dimensions[column_dimension].width = column_width

            col = column_index_from_string(column_dimension)
            
            cell = sheet.cell(row=row, column=col)
            cell.value = key
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        
        
        # # Add title
        # sheet['A1'] = "Summary of All Provinces"
        # sheet['A1'].font = Font(bold=True, size=16)
        # sheet.merge_cells('A1:H1')
        
        # # Add total cost
        # sheet['A2'] = "Total Project Cost:"
        # sheet['A2'].font = Font(bold=True)
        # sheet['B2'] = self.cost_data.get("total_cost", 0)
        # sheet['B2'].number_format = '#,##0 "VND"'
        # sheet.merge_cells('B2:H2')
        
        # # Get provinces
        # provinces = self.cost_data.get("provinces", {})
        # province_names = list(provinces.keys())
        
        # # Add province links section
        # row = 4
        # sheet.cell(row=row, column=1).value = "Province Cost Sheets:"
        # sheet.cell(row=row, column=1).font = Font(bold=True)
        # row += 1
        
        # # Add hyperlinks to each province sheet
        # for province in province_names:
        #     cell = sheet.cell(row=row, column=1)
        #     cell.value = f"Go to {province}"
        #     # Create internal hyperlink to the province sheet
        #     cell.hyperlink = f"#{province}!A1"
        #     cell.font = Font(color="0000FF", underline="single")
            
        #     # Add province total cost
        #     province_cost = provinces[province].get("total_cost", 0)
        #     sheet.cell(row=row, column=2).value = province_cost
        #     sheet.cell(row=row, column=2).number_format = '#,##0 "VND"'
            
        #     row += 1
        
        # row += 2  # Add spacing
        
        # Add headers for the main data table
        # row = 2

        # sheet[f'A{row}'] = "ESTIMATE COST"
        # sheet.merge_cells('A{row}:H{row}')

        # row = 9

        # headers = [
        #     "Subtitle / Component", 
        #     "Code", 
        #     "Unit",
        #     "Target Audience",
        #     "Qty",  # Moved before Unit Cost
        #     "Unit Cost (VND)",
        #     "Total Cost (VND)",
        #     "Note"
        # ]
        
        # header_row = row

        # for col, header in enumerate(headers, 1):
        #     cell = sheet.cell(row=header_row, column=col)
        #     cell.value = header
        #     cell.font = Font(bold=True)
        #     cell.alignment = Alignment(horizontal='center', vertical='center')
        #     cell.border = Border(
        #         left=Side(style='thin'),
        #         right=Side(style='thin'),
        #         top=Side(style='thin'),
        #         bottom=Side(style='thin')
        #     )
        
        # row = header_row + 1  # Start of data rows
        
        # # Get a combined hierarchy of all provinces
        # combined_hierarchy = self.get_combined_hierarchy(provinces)
        
        # # Add hierarchical data with totals
        # self.add_combined_hierarchy_totals_to_sheet(sheet, combined_hierarchy, row, 0, provinces)
        
        return sheet

    # def create_summary_tab(self):
    #     """Create a summary tab showing total costs for all provinces."""
    #     widget = QWidget()
    #     layout = QVBoxLayout(widget)
        
    #     # Create tree widget for summary
    #     tree = QTreeWidget()
    #     tree.setHeaderLabels(["Province / Component", "Cost (VND)"])
    #     tree.setAlternatingRowColors(True)
    #     tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
    #     tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        
    #     # Add provinces
    #     provinces = self.cost_data.get("provinces", {})
    #     for province, province_data in provinces.items():
    #         item = QTreeWidgetItem([province, f"{province_data.get('total_cost', 0):,.0f}"])
    #         item.setTextAlignment(1, Qt.AlignRight | Qt.AlignVCenter)
    #         tree.addTopLevelItem(item)
        
    #     # Add total
    #     total_item = QTreeWidgetItem(["Total Project Cost", f"{self.cost_data.get('total_cost', 0):,.0f}"])
    #     total_item.setTextAlignment(1, Qt.AlignRight | Qt.AlignVCenter)
    #     total_item.setBackground(0, QColor("#ECF0F1"))
    #     total_item.setBackground(1, QColor("#ECF0F1"))
    #     font = total_item.font(0)
    #     font.setBold(True)
    #     total_item.setFont(0, font)
    #     total_item.setFont(1, font)
    #     tree.addTopLevelItem(total_item)
        
    #     layout.addWidget(tree)
    #     return widget
        
    # def create_province_tab(self, province, province_data):
    #     """Create a tab for a specific province."""
    #     widget = QWidget()
    #     layout = QVBoxLayout(widget)
        
    #     # Province total cost
    #     # total_cost = province_data.get("total_cost", 0)
    #     # total_label = QLabel(f"Total Cost for {province}: {total_cost:,.0f} VND")
    #     # total_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 10px;")
    #     # layout.addWidget(total_label)
        
    #     # Create tree widget for costs
    #     tree = QTreeWidget()
    #     tree.setHeaderLabels([
    #         "Subtitle / Component", 
    #         "Code", 
    #         "Unit",
    #         "Target Audience", 
    #         "Unit Cost (VND)", 
    #         "Qty", 
    #         "Total Cost (VND)"
    #     ])
    #     tree.setAlternatingRowColors(True)
        
    #     # Set column widths
    #     tree.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Subtitle / Component
    #     tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Code
    #     tree.header().setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Unit
    #     tree.header().setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Target Audience
    #     tree.header().setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Unit Cost
    #     tree.header().setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Qty
    #     tree.header().setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Total Cost
        
    #     # Set minimum column widths
    #     tree.setColumnWidth(1, 80)  # Code
    #     tree.setColumnWidth(2, 80)  # Unit
    #     tree.setColumnWidth(3, 150)  # Target Audience
    #     tree.setColumnWidth(4, 120)  # Unit Cost
    #     tree.setColumnWidth(5, 100)  # Qty
    #     tree.setColumnWidth(6, 120)  # Total Cost
        
    #     # Add cost hierarchy
    #     self.add_cost_hierarchy(tree, province_data.get("hierarchy", {}))
        
    #     layout.addWidget(tree)
        
    #     # Expand top-level items
    #     for i in range(tree.topLevelItemCount()):
    #         tree.topLevelItem(i).setExpanded(True)
        
    #     # Enable custom tooltip handling
    #     tree.setMouseTracking(True)
        
    #     return widget
    
    

    # def add_cost_hierarchy(self, tree, hierarchy):
    #     """Add improved cost hierarchy to tree widget."""
    #     for subtitle, data in hierarchy.items():
    #         # Create item for this subtitle
    #         subtitle_item = QTreeWidgetItem([subtitle])
    #         subtitle_item.setTextAlignment(0, Qt.AlignLeft | Qt.AlignVCenter)
            
    #         # Set total cost in the last column
    #         subtitle_item.setText(6, f"{data.get('cost', 0):,.0f}")
    #         subtitle_item.setTextAlignment(6, Qt.AlignRight | Qt.AlignVCenter)
            
    #         # Set subtitle styling
    #         font = subtitle_item.font(0)
    #         font.setBold(True)
    #         subtitle_item.setFont(0, font)
    #         subtitle_item.setFont(6, font)
    #         subtitle_item.setBackground(0, QColor("#F2F2F2"))
    #         subtitle_item.setBackground(6, QColor("#F2F2F2"))
            
    #         tree.addTopLevelItem(subtitle_item)
            
    #         # Add elements
    #         for element in data.get("elements", []):
    #             code = element.get("code", "")
    #             unit = element.get("unit", "")
    #             target_audience = element.get("target_audience", "")
    #             base_cost = element.get("base_cost", 0)
    #             element_cost = element.get("element_cost", 0)
    #             element_cost_formula = element.get("element_cost_formula", "")
    #             coefficient = element.get("coefficient", 1.0)
    #             coefficient_formula = element.get("coefficient_formula", "")
    #             total_cost = element.get("total_cost", 0)
    #             level = element.get("level", "")  # Get level if it exists
    #             element_name = element.get("name", "")  # Get element name if it exists
    #             is_additional_cost = element.get("is_additional_cost", False)  # Check if this is an additional cost
                
    #             # Convert target_audience to string if it's a dictionary
    #             target_audience_str = self._get_target_audience_string(target_audience)
                
    #             # Create element item
    #             element_item = QTreeWidgetItem()
                
    #             # Set element properties
    #             # Add level to element name if it exists
    #             if not element_name:
    #                 element_name = subtitle
    #             if level:
    #                 element_name = f"{subtitle} ({level})"
    #             elif target_audience_str:
    #                 element_name = f"{subtitle} - {target_audience_str}"

    #             element_item.setText(0, element_name)
    #             element_item.setText(1, str(code))
    #             element_item.setText(2, unit)
    #             element_item.setText(3, target_audience_str)
    #             element_item.setText(4, f"{element_cost:,.0f}")
    #             element_item.setText(5, f"{coefficient:.1f}")
    #             element_item.setText(6, f"{total_cost:,.0f}")
                
    #             # Set text alignment
    #             element_item.setTextAlignment(1, Qt.AlignCenter)
    #             element_item.setTextAlignment(2, Qt.AlignCenter)
    #             element_item.setTextAlignment(3, Qt.AlignCenter)
    #             element_item.setTextAlignment(4, Qt.AlignRight | Qt.AlignVCenter)
    #             element_item.setTextAlignment(5, Qt.AlignRight | Qt.AlignVCenter)
    #             element_item.setTextAlignment(6, Qt.AlignRight | Qt.AlignVCenter)
                
    #             # Apply special formatting for additional costs
    #             if is_additional_cost:
    #                 # Make additional costs bold and add a light background color
    #                 for col in range(7):
    #                     font = element_item.font(col)
    #                     font.setBold(True)
    #                     element_item.setFont(col, font)
    #                     # Light blue background to distinguish additional costs
    #                     element_item.setBackground(col, QColor("#E3F2FD"))
                
    #             # Set tooltips with formulas
    #             element_item.setToolTip(4, element_cost_formula)  # Element cost formula on hover
    #             element_item.setToolTip(5, coefficient_formula)   # Qty formula on hover
                
    #             subtitle_item.addChild(element_item)
            
    #         # Add children subtitles recursively
    #         self.add_subtitle_children(subtitle_item, data.get("children", {}))

    # def add_subtitle_children(self, parent_item, children):
    #     """Add improved subtitle children to tree item."""
    #     for subtitle, data in children.items():
    #         # Create item for this subtitle
    #         subtitle_item = QTreeWidgetItem()
    #         subtitle_item.setText(0, subtitle)
    #         subtitle_item.setText(6, f"{data.get('cost', 0):,.0f}")
            
    #         # Set text alignment
    #         subtitle_item.setTextAlignment(0, Qt.AlignLeft | Qt.AlignVCenter)
    #         subtitle_item.setTextAlignment(6, Qt.AlignRight | Qt.AlignVCenter)
            
    #         # Set subtitle styling
    #         font = subtitle_item.font(0)
    #         font.setBold(True)
    #         subtitle_item.setFont(0, font)
    #         subtitle_item.setFont(6, font)
            
    #         # Set a lighter background to differentiate child subtitles
    #         subtitle_item.setBackground(0, QColor("#F8F8F8"))
    #         subtitle_item.setBackground(6, QColor("#F8F8F8"))
            
    #         parent_item.addChild(subtitle_item)
            
    #         # Add elements
    #         for element in data.get("elements", []):
    #             code = element.get("code", "")
                
    #             # Fix for nan values in unit
    #             unit = element.get("unit", "")
    #             if isinstance(unit, float) and (unit != unit or pd.isna(unit)):  # Check for nan
    #                 unit = ""
                    
    #             target_audience = element.get("target_audience", "")
    #             if isinstance(target_audience, float) and (target_audience != target_audience or pd.isna(target_audience)):
    #                 target_audience = ""
                
    #             # Convert target_audience to string if it's a dictionary
    #             target_audience_str = self._get_target_audience_string(target_audience)
                    
    #             base_cost = element.get("base_cost", 0)
    #             element_cost = element.get("element_cost", 0)
    #             element_cost_formula = element.get("element_cost_formula", "")
    #             coefficient = element.get("coefficient", 1.0)
    #             coefficient_formula = element.get("coefficient_formula", "")
    #             total_cost = element.get("total_cost", 0)
    #             level = element.get("level", "")  # Get level if it exists
    #             element_name = element.get("name", "")  # Get element name if it exists
    #             is_additional_cost = element.get("is_additional_cost", False)  # Check if this is an additional cost

    #             # Create element item
    #             element_item = QTreeWidgetItem()
                
    #             # Set element properties
    #             # Add level to element name if it exists
    #             if not element_name:
    #                 element_name = subtitle
    #             if level:
    #                 element_name = f"{element_name} ({level})"
    #             elif target_audience_str:
    #                 element_name = f"{element_name} - {target_audience_str}"

    #             element_item.setText(0, element_name)
    #             element_item.setText(1, str(code))
    #             element_item.setText(2, str(unit))  # Ensure it's a string
    #             element_item.setText(3, str(target_audience_str))  # Ensure it's a string
    #             element_item.setText(4, f"{element_cost:,.0f}")
    #             element_item.setText(5, f"{coefficient:.1f}")
    #             element_item.setText(6, f"{total_cost:,.0f}")
                
    #             # Set text alignment
    #             element_item.setTextAlignment(1, Qt.AlignCenter)
    #             element_item.setTextAlignment(2, Qt.AlignCenter)
    #             element_item.setTextAlignment(3, Qt.AlignCenter)
    #             element_item.setTextAlignment(4, Qt.AlignRight | Qt.AlignVCenter)
    #             element_item.setTextAlignment(5, Qt.AlignRight | Qt.AlignVCenter)
    #             element_item.setTextAlignment(6, Qt.AlignRight | Qt.AlignVCenter)
                
    #             # Apply special formatting for additional costs
    #             if is_additional_cost:
    #                 # Make additional costs bold and add a light background color
    #                 for col in range(7):
    #                     font = element_item.font(col)
    #                     font.setBold(True)
    #                     element_item.setFont(col, font)
    #                     # Light blue background to distinguish additional costs
    #                     element_item.setBackground(col, QColor("#E3F2FD"))
                
    #             # Set tooltips with formulas
    #             element_item.setToolTip(4, element_cost_formula)  # Element cost formula on hover
    #             element_item.setToolTip(5, coefficient_formula)   # Qty formula on hover
                
    #             subtitle_item.addChild(element_item)
            
    #         # Add children subtitles recursively
    #         self.add_subtitle_children(subtitle_item, data.get("children", {}))
    
    
    # def create_project_info_sheet(self, wb, project_model=None):
    #     """Create and populate the Project Information sheet."""
    #     # Create sheet
    #     sheet = wb.create_sheet("Project Information")
        
    #     # Set column widths
    #     sheet.column_dimensions['A'].width = 30
    #     sheet.column_dimensions['B'].width = 50
        
    #     # Add title
    #     sheet['A1'] = "Project Information"
    #     sheet['A1'].font = Font(bold=True, size=16)
    #     sheet.merge_cells('A1:B1')
        
    #     row = 3
        
    #     # Add project information
    #     sheet[f'A{row}'] = "Project Information"
    #     sheet[f'A{row}'].font = Font(bold=True)
    #     row += 1
        
    #     # If project_model is available, extract information from it
    #     if project_model:
    #         # Extract general information
    #         general = project_model.general
            
    #         # Add project details
    #         fields = [
    #             ("Project Name", general.get("project_name", "")),
    #             ("Project Type", general.get("project_type", "")),
    #             ("Internal Job", general.get("internal_job", "")),
    #             ("Symphony", general.get("symphony", "")),
    #             ("Sampling Method", general.get("sampling_method", "")),
    #             ("Responsibility Classification", general.get("resp_classification", "")),
    #             ("Interview Length", f"{general.get('interview_length', 0)} minutes"),
    #             ("Questionnaire Length", general.get("questionnaire_length", 0)),
    #             ("Open-ended Questions", general.get("open_ended_count", 0)),
    #             ("Interview Methods", ", ".join(general.get("interview_methods", []))),
    #             ("Service Line", general.get("service_line", "")),
    #             ("Industries", ", ".join(general.get("industries", []))),
    #             ("Provinces", ", ".join(general.get("provinces", []))),
    #             ("Sample Types", ", ".join(general.get("sample_types", [])))
    #         ]
            
    #         for field, value in fields:
    #             sheet[f'A{row}'] = field + ":"
    #             sheet[f'B{row}'] = value
    #             row += 1
            
    #         # Update Target Audiences section
    #         if project_model.general.get("target_audiences"):
    #             row += 2
    #             sheet[f'A{row}'] = "Target Audiences"
    #             sheet[f'A{row}'].font = Font(bold=True)
    #             row += 1
                
    #             # Add header for target audiences table
    #             sheet[f'A{row}'] = "Target Audience"
    #             sheet[f'B{row}'] = "Representative Price Growth Rate (%)"
                
    #             # Style headers
    #             for col in ['A', 'B']:
    #                 cell = sheet[f'{col}{row}']
    #                 cell.font = Font(bold=True)
    #                 cell.alignment = Alignment(horizontal='center')
    #                 cell.border = Border(
    #                     left=Side(style='thin'), right=Side(style='thin'),
    #                     top=Side(style='thin'), bottom=Side(style='thin')
    #                 )
                
    #             row += 1
                
    #             # Check if there are any "All Audiences" elements
    #             has_aggregated_elements = False
    #             for province_name, province_data in self.cost_data.get("provinces", {}).items():
    #                 hierarchy = province_data.get("hierarchy", {})
    #                 if self._check_for_aggregated_elements(hierarchy):
    #                     has_aggregated_elements = True
    #                     break

    #             # Process target audiences (now list of dictionaries)
    #             provinces = project_model.general.get("provinces", [])
    #             target_audiences = project_model.general.get("target_audiences", [])
    #             sample_types = project_model.general.get("sample_types", [])
                
    #             # Create list of audiences to show
    #             audiences_to_show = []
                
    #             # Add each target audience dictionary (only if unique combination)
    #             added_combinations = set()
    #             for audience_dict in target_audiences:
    #                 # Create a unique key based on the audience attributes
    #                 unique_key = (
    #                     audience_dict.get('name', ''),
    #                     tuple(audience_dict.get('age_range', [])),
    #                     tuple(audience_dict.get('income_range', [])),
    #                     audience_dict.get('incident_rate', 0),
    #                     audience_dict.get('complexity', '')
    #                 )
                    
    #                 # Only add if this combination hasn't been seen before
    #                 if unique_key not in added_combinations:
    #                     added_combinations.add(unique_key)
    #                     audiences_to_show.append(audience_dict)
                
    #             for audience in audiences_to_show:
    #                 # Handle audience dictionary case
    #                 audience_display_name = self._get_audience_display_name(audience)
    #                 representative_growth_rate = self._find_representative_price_growth(
    #                     audience, provinces, project_model.samples
    #                 )
                    
    #                 # Add target audience row
    #                 sheet[f'A{row}'] = audience_display_name
    #                 sheet[f'B{row}'] = representative_growth_rate
                    
    #                 # Style cells
    #                 for col in ['A', 'B']:
    #                     cell = sheet[f'{col}{row}']
    #                     cell.border = Border(
    #                         left=Side(style='thin'), right=Side(style='thin'),
    #                         top=Side(style='thin'), bottom=Side(style='thin')
    #                     )
                    
    #                 sheet[f'B{row}'].number_format = '0.0'
    #                 sheet[f'B{row}'].alignment = Alignment(horizontal='right')
                    
    #                 row += 1
                
    #             # Add HUT information if applicable
    #             if general.get("hut_test_products", 0) > 0:
    #                 row += 1
    #                 sheet[f'A{row}'] = "HUT (Home Use Test)"
    #                 sheet[f'A{row}'].font = Font(bold=True)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "Test Products:"
    #                 sheet[f'B{row}'] = general.get("hut_test_products", 0)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "Usage Duration (days):"
    #                 sheet[f'B{row}'] = general.get("hut_usage_duration", 0)
    #                 row += 1
                
    #             # Add CLT information if applicable
    #             if general.get("clt_test_products", 0) > 0:
    #                 row += 1
    #                 sheet[f'A{row}'] = "CLT (Central Location Test)"
    #                 sheet[f'A{row}'].font = Font(bold=True)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "Test Products:"
    #                 sheet[f'B{row}'] = general.get("clt_test_products", 0)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "Respondent Visits:"
    #                 sheet[f'B{row}'] = general.get("clt_respondent_visits", 0)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "Sample Size Target per Day:"
    #                 sheet[f'B{row}'] = general.get("clt_sample_size_per_day", 0)
    #                 row += 1
                
    #             # Add Printer information if applicable
    #             if any(general.get(field, 0) > 0 for field in ["showcard_page_count", "dropcard_page_count", "color_page_count", "laminated_page_count"]):
    #                 row += 1
    #                 sheet[f'A{row}'] = "Printer"
    #                 sheet[f'A{row}'].font = Font(bold=True)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "SHOWCARD Page Count:"
    #                 sheet[f'B{row}'] = general.get("showcard_page_count", 0)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "DROPCARD Page Count:"
    #                 sheet[f'B{row}'] = general.get("dropcard_page_count", 0)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "COLOR Page Count:"
    #                 sheet[f'B{row}'] = general.get("color_page_count", 0)
    #                 row += 1
                    
    #                 sheet[f'A{row}'] = "Laminated Page Count:"
    #                 sheet[f'B{row}'] = general.get("laminated_page_count", 0)
    #                 row += 1
    #     else:
    #         # If project_model is not available, add a placeholder
    #         sheet[f'A{row}'] = "Project model not available"
    #         row += 1
            
    #     # Add section header for Samples information
    #     if hasattr(project_model, 'samples') and project_model.samples:
    #         row += 2
    #         sheet[f'A{row}'] = "Samples Information"
    #         sheet[f'A{row}'].font = Font(bold=True)
    #         row += 1
            
    #         # Add header for the sample summary table
    #         headers = ["Province", "Total Sample Size", "Comments"]
            
    #         for col_idx, header in enumerate(headers):
    #             col_letter = chr(65 + col_idx)  # A, B, C
    #             cell = sheet[f'{col_letter}{row}']
    #             cell.value = header
    #             cell.font = Font(bold=True)
    #             cell.alignment = Alignment(horizontal='center')
    #             cell.border = Border(
    #                 left=Side(style='thin'), right=Side(style='thin'),
    #                 top=Side(style='thin'), bottom=Side(style='thin')
    #             )
                
    #             # Set column widths
    #             if col_idx == 0:  # Province
    #                 sheet.column_dimensions[col_letter].width = 20
    #             elif col_idx == 1:  # Total Sample Size
    #                 sheet.column_dimensions[col_letter].width = 15
    #             else:  # Comments
    #                 sheet.column_dimensions[col_letter].width = 50
            
    #         row += 1
            
    #         # Process samples with new structure (list of sample entries per province)
    #         provinces = project_model.general.get("provinces", [])
    #         target_audiences = project_model.general.get("target_audiences", [])
            
    #         for province in provinces:
    #             total_sample_size = 0
    #             comments = []
                
    #             # Calculate total sample size and collect comments from list structure
    #             if province in project_model.samples:
    #                 for sample_entry_key, sample_entry_value in project_model.samples[province].items():
    #                     sample_size = sample_entry_value.get("sample_size", 0)
    #                     total_sample_size += sample_size
                        
    #                     # Collect comments related to sample size and price growth
    #                     if "comment" in sample_entry_value:
    #                         comment_dict = sample_entry_value["comment"]
    #                         audience_name = sample_entry_value.get("name", "")
    #                         sample_type = sample_entry_value.get("sample_type", "")
                            
    #                         if "sample_size" in comment_dict:
    #                             comments.append(f"{audience_name} - {sample_type} - Sample Size: {comment_dict['sample_size']}")
    #                         if "price_growth" in comment_dict:
    #                             comments.append(f"{audience_name} - {sample_type} - Price Growth: {comment_dict['price_growth']}")
                
    #             # Add province row
    #             sheet[f'A{row}'] = province
    #             sheet[f'B{row}'] = total_sample_size
    #             sheet[f'C{row}'] = "\n".join(comments)
                
    #             # Style cells
    #             for col in ['A', 'B', 'C']:
    #                 cell = sheet[f'{col}{row}']
    #                 cell.border = Border(
    #                     left=Side(style='thin'), right=Side(style='thin'),
    #                     top=Side(style='thin'), bottom=Side(style='thin')
    #                 )
                
    #             sheet[f'B{row}'].alignment = Alignment(horizontal='right')
    #             sheet[f'C{row}'].alignment = Alignment(wrap_text=True)
                
    #             row += 1
        
    #     # Add QC Method information
    #     if project_model and hasattr(project_model, 'qc_methods') and project_model.qc_methods:
    #         row += 2
    #         sheet[f'A{row}'] = "QC Methods"
    #         sheet[f'A{row}'].font = Font(bold=True)
    #         row += 1
            
    #         # Add header for QC methods table
    #         sheet[f'A{row}'] = "Team"
    #         sheet[f'B{row}'] = "Method"
    #         sheet[f'C{row}'] = "Rate (%)"
            
    #         # Style headers
    #         for col in ['A', 'B', 'C']:
    #             cell = sheet[f'{col}{row}']
    #             cell.font = Font(bold=True)
    #             cell.alignment = Alignment(horizontal='center')
    #             cell.border = Border(
    #                 left=Side(style='thin'), right=Side(style='thin'),
    #                 top=Side(style='thin'), bottom=Side(style='thin')
    #             )
            
    #         row += 1
            
    #         # Add QC methods
    #         for qc_method in project_model.qc_methods:
    #             sheet[f'A{row}'] = qc_method.get("team", "")
    #             sheet[f'B{row}'] = qc_method.get("method", "")
    #             sheet[f'C{row}'] = qc_method.get("rate", 0)
                
    #             # Style cells
    #             for col in ['A', 'B', 'C']:
    #                 cell = sheet[f'{col}{row}']
    #                 cell.border = Border(
    #                     left=Side(style='thin'), right=Side(style='thin'),
    #                     top=Side(style='thin'), bottom=Side(style='thin')
    #                 )
                
    #             sheet[f'C{row}'].alignment = Alignment(horizontal='right')
                
    #             row += 1
        
    #     # Add section header for cost summary
    #     row += 2
    #     sheet[f'A{row}'] = "Cost Summary"
    #     sheet[f'A{row}'].font = Font(bold=True)
    #     row += 1
        
    #     # Add total project cost
    #     sheet[f'A{row}'] = "Total Project Cost:"
    #     sheet[f'B{row}'] = self.cost_data.get("total_cost", 0)
    #     sheet[f'B{row}'].number_format = '#,##0 "VND"'
    #     row += 1
        
    #     # Add costs by province
    #     provinces = self.cost_data.get("provinces", {})
    #     for province, province_data in provinces.items():
    #         sheet[f'A{row}'] = f"{province} Cost:"
    #         sheet[f'B{row}'] = province_data.get("total_cost", 0)
    #         sheet[f'B{row}'].number_format = '#,##0 "VND"'
    #         row += 1

    # def _get_audience_display_name(self, audience_dict):
    #     """Get display name for an audience dictionary."""
    #     if isinstance(audience_dict, str):
    #         return audience_dict
        
    #     name = audience_dict.get('name', '')
    #     age_range = audience_dict.get('age_range', [])
    #     income_range = audience_dict.get('income_range', [])
    #     incident_rate = audience_dict.get('incident_rate', 0)
    #     complexity = audience_dict.get('complexity', '')
        
    #     return f"{name} - Age {age_range} - Income {income_range} - IR {incident_rate} - Complexity: {complexity}"

    # def _find_representative_price_growth(self, audience_dict, provinces, samples):
    #     """Find representative price growth rate for an audience dictionary."""
    #     representative_growth_rate = 0.0
        
    #     # Search through the new list-based samples structure
    #     for province in provinces:
    #         if province in samples:
    #             for sample_entry_key, sample_entry_value in samples[province].items():
    #                 # Check if this sample entry matches the audience criteria
    #                 if self._audience_matches(audience_dict, sample_entry_value):
    #                     representative_growth_rate = sample_entry_value.get("price_growth", 0.0)
    #                     if representative_growth_rate > 0:
    #                         return representative_growth_rate
        
    #     return representative_growth_rate

    # def _audience_matches(self, audience_dict, sample_entry):
    #     """Check if a sample entry matches the audience criteria."""
    #     return (
    #         sample_entry.get('name') == audience_dict.get('name') and
    #         sample_entry.get('age_range') == audience_dict.get('age_range') and
    #         sample_entry.get('income_range') == audience_dict.get('income_range') and
    #         sample_entry.get('incident_rate') == audience_dict.get('incident_rate') and
    #         sample_entry.get('complexity') == audience_dict.get('complexity')
    #     )
    
    def create_summary_sheet(self, wb, project_model=None):
        """Create and populate the Summary sheet with total costs across all provinces."""
        # Create sheet
        sheet = wb.create_sheet("Summary")
        
        # Set column widths for initial columns
        sheet.column_dimensions['A'].width = 40  # Subtitle / Component
        sheet.column_dimensions['B'].width = 15  # Code
        sheet.column_dimensions['C'].width = 10  # Unit
        sheet.column_dimensions['D'].width = 20  # Target Audience
        sheet.column_dimensions['E'].width = 15  # Qty (moved before Unit Cost)
        sheet.column_dimensions['F'].width = 20  # Unit Cost
        sheet.column_dimensions['G'].width = 20  # Total Cost
        sheet.column_dimensions['H'].width = 50  # Note
        
        # Add title
        sheet['A1'] = "Summary of All Provinces"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:H1')
        
        # Add total cost
        sheet['A2'] = "Total Project Cost:"
        sheet['A2'].font = Font(bold=True)
        sheet['B2'] = self.cost_data.get("total_cost", 0)
        sheet['B2'].number_format = '#,##0 "VND"'
        sheet.merge_cells('B2:H2')
        
        # Get provinces
        provinces = self.cost_data.get("provinces", {})
        province_names = list(provinces.keys())
        
        # Add province links section
        row = 4
        sheet.cell(row=row, column=1).value = "Province Cost Sheets:"
        sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        # Add hyperlinks to each province sheet
        for province in province_names:
            cell = sheet.cell(row=row, column=1)
            cell.value = f"Go to {province}"
            # Create internal hyperlink to the province sheet
            cell.hyperlink = f"#{province}!A1"
            cell.font = Font(color="0000FF", underline="single")
            
            # Add province total cost
            province_cost = provinces[province].get("total_cost", 0)
            sheet.cell(row=row, column=2).value = province_cost
            sheet.cell(row=row, column=2).number_format = '#,##0 "VND"'
            
            row += 1
        
        row += 2  # Add spacing
        
        # Add headers for the main data table
        headers = [
            "Subtitle / Component", 
            "Code", 
            "Unit",
            "Target Audience",
            "Qty",  # Moved before Unit Cost
            "Unit Cost (VND)",
            "Total Cost (VND)",
            "Note"
        ]
        
        header_row = row
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row = header_row + 1  # Start of data rows
        
        # Get a combined hierarchy of all provinces
        combined_hierarchy = self.get_combined_hierarchy(provinces)
        
        # Add hierarchical data with totals
        self.add_combined_hierarchy_totals_to_sheet(sheet, combined_hierarchy, row, 0, provinces)
        
        return sheet

    def add_combined_hierarchy_totals_to_sheet(self, sheet, hierarchy, start_row, level, provinces, parent_path=None):
        """
        Add combined hierarchical cost data with totals to a sheet.
        
        Args:
            sheet: Excel worksheet
            hierarchy: The combined hierarchy to add
            start_row: Starting row index
            level: Current hierarchy level (0 for top level)
            provinces: Dictionary of province data
            parent_path: List of parent subtitles (for tracking nested path)
        """
        row = start_row
        
        # Initialize parent path if needed
        if parent_path is None:
            parent_path = []
        
        for subtitle, data in hierarchy.items():
            # Current path includes all parent subtitles plus current subtitle
            current_path = parent_path + [subtitle]
            
            # Skip adding subtitle rows - go directly to elements
            
            # Process elements - group by code, target audience, AND subtitle path
            elements_by_unit_cost = {}

            for element in data.get("elements", []):
                code = element.get("code", "")
                target_audience = element.get("target_audience", "")
                level_value = element.get("level", "")
                
                # Convert target_audience to string if it's a dictionary
                target_audience_str = self._get_target_audience_string(target_audience)
                
                # Include the current path in the key for uniqueness
                path_key = tuple(current_path)
                
                # Find this element in all provinces and group by unit cost
                for province_name, province_data in provinces.items():
                    # For "All Audiences" elements, search without specific target audience
                    search_target_audience = None if target_audience_str == "All Audiences" else target_audience
                    
                    province_element = self.find_element_in_hierarchy(
                        province_data.get("hierarchy", {}), 
                        code,
                        search_target_audience,
                        list(path_key),
                        level_value
                    )
                    
                    if province_element:
                        element_cost = province_element.get("element_cost", 0)
                        
                        # Create key using string representation of target_audience
                        if level_value:
                            key = (code, target_audience_str, path_key, level_value, element_cost)
                        else:
                            key = (code, target_audience_str, path_key, element_cost)
                        
                        if key not in elements_by_unit_cost:
                            elements_by_unit_cost[key] = {
                                "element": element,
                                "provinces": {},
                                "unit_cost": element_cost
                            }
                        
                        elements_by_unit_cost[key]["provinces"][province_name] = province_element
            
            # Add each unique element with same unit cost
            for key, grouped_data in elements_by_unit_cost.items():
                element = grouped_data["element"]
                province_elements = grouped_data["provinces"]
                unit_cost = grouped_data["unit_cost"]
                
                code = element.get("code", "")
                target_audience = element.get("target_audience", "")
                level_value = element.get("level", "")
                
                # Convert target_audience to string for display
                target_audience_str = self._get_target_audience_string(target_audience)
                
                # Add element details
                element_name = "  " * (level + 1) + subtitle
                if level_value:
                    element_name += f" ({level_value})"
                    
                sheet.cell(row=row, column=1).value = element_name
                sheet.cell(row=row, column=2).value = code
                sheet.cell(row=row, column=3).value = element.get("unit", "")
                sheet.cell(row=row, column=4).value = target_audience_str
                
                # Calculate totals for provinces with same unit cost
                total_qty = 0
                total_cost = 0
                notes_dict = {}  # Store unique notes with province info
                
                for province_name, province_element in province_elements.items():
                    qty = province_element.get("coefficient", 0)
                    element_total = province_element.get("total_cost", 0)
                    
                    total_qty += qty
                    total_cost += element_total
                    
                    # Collect notes
                    element_cost_formula = province_element.get("element_cost_formula", "")
                    coefficient_formula = province_element.get("coefficient_formula", "")
                    
                    # Create note key to identify unique formulas
                    note_key = (element_cost_formula, coefficient_formula)
                    if note_key not in notes_dict:
                        notes_dict[note_key] = []
                    notes_dict[note_key].append(province_name)
                
                # Add totals to sheet
                # Column 5 (E) - Qty
                cell = sheet.cell(row=row, column=5)
                cell.value = total_qty
                cell.number_format = '#,##0.0'
                
                # Column 6 (F) - Unit Cost
                cell = sheet.cell(row=row, column=6)
                cell.value = unit_cost
                cell.number_format = '#,##0'
                
                # Column 7 (G) - Total Cost
                cell = sheet.cell(row=row, column=7)
                cell.value = total_cost
                cell.number_format = '#,##0'
                
                # Column 8 (H) - Note (combine notes from provinces with same unit cost)
                note_parts = []
                for (element_cost_formula, coefficient_formula), province_list in notes_dict.items():
                    if element_cost_formula or coefficient_formula:
                        province_str = ", ".join(province_list)
                        note_part = f"[{province_str}]\n"
                        if element_cost_formula:
                            note_part += f"Unit Cost: {element_cost_formula}\n"
                        if coefficient_formula:
                            note_part += f"Qty: {coefficient_formula}"
                        note_parts.append(note_part)
                
                if note_parts:
                    note_text = "\n\n".join(note_parts)
                    sheet.cell(row=row, column=8).value = note_text
                
                # Add borders and alignment for all cells in this row
                for col in range(1, 9):
                    cell = sheet.cell(row=row, column=col)
                    if not cell.value:  # Empty cell
                        cell.value = "-"  # Add placeholder
                    
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Set alignment based on column type
                    if col in [5, 6, 7]:  # Numeric columns (Qty, Unit Cost, Total Cost)
                        cell.alignment = Alignment(horizontal='right')
                    elif col == 8:  # Note column
                        cell.alignment = Alignment(wrap_text=True)
                
                row += 1
            
            # Add children subtitles recursively
            children = data.get("children", {})
            if children:
                # Pass current path as parent path for children
                row = self.add_combined_hierarchy_totals_to_sheet(
                    sheet, children, row, level + 1, provinces, current_path
                )
        
        return row

    def add_hierarchy_to_sheet(self, sheet, hierarchy, start_row, level):
        """Add hierarchical cost data to a sheet."""
        row = start_row
        
        for subtitle, data in hierarchy.items():
            # Skip adding subtitle rows - go directly to elements
            
            # Add elements
            for element in data.get("elements", []):
                target_audience = element.get("target_audience", "")
                level_value = element.get("level", "")
                
                # Convert target_audience to string if it's a dictionary
                target_audience_str = self._get_target_audience_string(target_audience)
                
                # Add element details
                element_name = "  " * (level + 1) + subtitle
                if level_value:
                    element_name += f" ({level_value})"
                    
                sheet.cell(row=row, column=1).value = element_name
                sheet.cell(row=row, column=2).value = element.get("code", "")
                sheet.cell(row=row, column=3).value = element.get("unit", "")
                sheet.cell(row=row, column=4).value = target_audience_str
                
                # Add element cost
                cell = sheet.cell(row=row, column=5)
                cell.value = element.get("element_cost", 0)
                cell.number_format = '#,##0'
                
                # Add coefficient
                cell = sheet.cell(row=row, column=6)
                cell.value = element.get("coefficient", 0)
                cell.number_format = '#,##0.0'
                
                # Add total cost
                cell = sheet.cell(row=row, column=7)
                cell.value = element.get("total_cost", 0)
                cell.number_format = '#,##0'
                
                # Add note with formulas
                element_cost_formula = element.get("element_cost_formula", "")
                coefficient_formula = element.get("coefficient_formula", "")
                note = f"Unit Cost: {element_cost_formula}\nQty: {coefficient_formula}"
                sheet.cell(row=row, column=8).value = note
                
                # Add borders and alignment
                for col in range(1, 9):
                    cell = sheet.cell(row=row, column=col)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
                    
                    # Align text
                    if col in (5, 6, 7):  # Numeric columns
                        cell.alignment = Alignment(horizontal='right')
                    elif col == 8:  # Note column
                        cell.alignment = Alignment(wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='left')
                
                row += 1
            
            # Add children subtitles recursively
            children = data.get("children", {})
            if children:
                row = self.add_hierarchy_to_sheet(sheet, children, row, level + 1)
        
        return row

    def find_element_in_hierarchy(self, hierarchy, code, target_audience=None, path=None, level=None):
        """Find an element with the given code and target audience in a hierarchy.
        
        Args:
            hierarchy: The hierarchy to search in
            code: The element code to find
            target_audience: The target audience to find (or None for any)
            path: The path of subtitles to follow (or None to search entire hierarchy)
            level: The level to match (or None for any)
            
        Returns:
            The element if found, None otherwise
        """
        # If path is provided, follow it to find the element
        if path and len(path) > 0:
            current_subtitle = path[0]
            
            # Check if this subtitle exists in the hierarchy
            if current_subtitle in hierarchy:
                if len(path) == 1:
                    # We're at the target subtitle, look for the element here
                    for element in hierarchy[current_subtitle].get("elements", []):
                        if element.get("code", "") == code:
                            # Match target audience if specified
                            element_audience = element.get("target_audience", "")
                            element_audience_str = self._get_target_audience_string(element_audience)
                            
                            if target_audience is None:
                                audience_match = True
                            elif element_audience_str == "All Audiences":
                                # "All Audiences" elements match any target audience search
                                audience_match = True
                            else:
                                # Compare with string representation of target audience
                                target_audience_str = self._get_target_audience_string(target_audience)
                                audience_match = element_audience_str == target_audience_str
                            
                            # Match level if specified
                            level_match = level is None or element.get("level", "") == level
                            
                            if audience_match and level_match:
                                return element  
                else:
                    # Still need to follow the path deeper
                    return self.find_element_in_hierarchy(
                        hierarchy[current_subtitle].get("children", {}),
                        code,
                        target_audience,
                        path[1:],
                        level
                    )
            
            # If we get here, the element wasn't found following the path
            return None
        
        # If no path is provided, search the entire hierarchy (original behavior)
        for subtitle, data in hierarchy.items():
            # Check elements at this level
            for element in data.get("elements", []):
                if element.get("code", "") == code:
                    # Match target audience if specified
                    element_audience = element.get("target_audience", "")
                    element_audience_str = self._get_target_audience_string(element_audience)
                    
                    if target_audience is None:
                        audience_match = True
                    elif element_audience_str == "All Audiences":
                        # "All Audiences" elements match any target audience search
                        audience_match = True
                    else:
                        # Compare with string representation of target audience
                        target_audience_str = self._get_target_audience_string(target_audience)
                        audience_match = element_audience_str == target_audience_str
                    
                    # Match level if specified
                    level_match = level is None or element.get("level", "") == level
                    
                    if audience_match and level_match:
                        return element
            
            # Check children recursively
            result = self.find_element_in_hierarchy(data.get("children", {}), code, target_audience, None, level)
            if result:
                return result
        
        return None

    def _get_target_audience_string(self, target_audience):
        """Convert target_audience to string representation for consistent handling.
        
        Args:
            target_audience: Can be a string, dictionary, or other type
            
        Returns:
            str: String representation of the target audience
        """
        if isinstance(target_audience, str):
            return target_audience
        elif isinstance(target_audience, dict):
            # Convert dictionary to string representation matching the format used elsewhere
            name = target_audience.get('name', '')
            age_range = target_audience.get('age_range', [])
            income_range = target_audience.get('income_range', [])
            incident_rate = target_audience.get('incident_rate', 0)
            complexity = target_audience.get('complexity', '')
            return f"{name} - Age {age_range} - Income {income_range} - IR {incident_rate} - Complexity: {complexity}"
        elif target_audience is None or target_audience == "":
            return ""
        else:
            # Handle any other types (float NaN, etc.)
            return str(target_audience)

    def get_combined_hierarchy(self, provinces):
        """Create a combined hierarchy from all provinces."""
        combined = {}
        
        for province_name, province_data in provinces.items():
            hierarchy = province_data.get("hierarchy", {})
            
            # Merge this province's hierarchy into the combined hierarchy
            self.merge_hierarchies(combined, hierarchy)
        
        return combined

    def merge_hierarchies(self, combined, hierarchy):
        """Merge a province hierarchy into the combined hierarchy."""
        for subtitle, data in hierarchy.items():
            # If this subtitle doesn't exist in combined, add it
            if subtitle not in combined:
                combined[subtitle] = {
                    "children": {},
                    "elements": [],
                    "cost": 0
                }
            
            # Merge elements (just add to the list - we'll process them later)
            combined[subtitle]["elements"].extend(data.get("elements", []))
            
            # Recursively merge children
            children = data.get("children", {})
            if children:
                if "children" not in combined[subtitle]:
                    combined[subtitle]["children"] = {}
                
                self.merge_hierarchies(combined[subtitle]["children"], children)

    def find_subtitle_in_hierarchy(self, hierarchy, subtitle_name):
        """Find a subtitle with the given name in a hierarchy.
        
        Args:
            hierarchy (dict): The hierarchy to search in
            subtitle_name (str): The subtitle name to find
            
        Returns:
            dict or None: The subtitle data if found, None otherwise
        """
        # Check if the subtitle exists at this level
        if subtitle_name in hierarchy:
            return hierarchy[subtitle_name]
        
        # If not, check in all children
        for _, data in hierarchy.items():
            result = self.find_subtitle_in_hierarchy(data.get("children", {}), subtitle_name)
            if result:
                return result
        
        return None
    
    def create_province_sheet(self, wb, province, province_data):
        """Create and populate a sheet for a province."""
        # Create sheet
        sheet = wb.create_sheet(province)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 40  # Subtitle / Component
        sheet.column_dimensions['B'].width = 15  # Code
        sheet.column_dimensions['C'].width = 10  # Unit
        sheet.column_dimensions['D'].width = 20  # Target Audience
        sheet.column_dimensions['E'].width = 20  # Unit Cost (VND)
        sheet.column_dimensions['F'].width = 15  # Qty
        sheet.column_dimensions['G'].width = 20  # Total Cost (VND)
        sheet.column_dimensions['H'].width = 50  # Note
        
        # Add title
        sheet['A1'] = f"{province} Cost Details"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:H1')
        
        # Add total cost
        sheet['A2'] = f"Total Cost for {province}:"
        sheet['A2'].font = Font(bold=True)
        sheet['B2'] = province_data.get("total_cost", 0)
        sheet['B2'].number_format = '#,##0 "VND"'
        sheet.merge_cells('B2:H2')
        
        # Add headers
        headers = [
            "Subtitle / Component", 
            "Code", 
            "Unit",
            "Target Audience", 
            "Unit Cost (VND)", 
            "Qty", 
            "Total Cost (VND)",
            "Note"
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Add border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = border
        
        row += 1
        
        # Add hierarchical cost data
        hierarchy = province_data.get("hierarchy", {})
        self.add_hierarchy_to_sheet(sheet, hierarchy, row, 0)

    def _check_for_aggregated_elements(self, hierarchy):
        """Check if hierarchy contains any 'All Audiences' elements."""
        for subtitle, data in hierarchy.items():
            # Check elements at this level
            for element in data.get("elements", []):
                if element.get("target_audience") == "All Audiences":
                    return True
            
            # Check children recursively
            if data.get("children") and self._check_for_aggregated_elements(data["children"]):
                return True
        
        return False